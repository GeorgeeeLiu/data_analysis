import pandas as pd
from openpyxl import load_workbook
import os
import xlsxwriter
import requests
import jsonpath
import json
import xlrd
import xlwt
from xlutils.copy import copy
from chaowan import getSeriesID, getSeriesInfo, getProductInfo


def excelAddSheet(dataframe, sheetname, excelWriter):
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheetname, index=None)
    excelWriter.close()


URL = 'https://api.imagefield.cn/trade/tags?limit=99999&offset=0&parentId=1&orderBy=followerCount'
re = requests.get(URL)
html = json.loads(re.text)
count = jsonpath.jsonpath(html, 'count')[0]

IP_id = jsonpath.jsonpath(html, '$..id')  # id
IP_name = jsonpath.jsonpath(html, '$..name')  # 名称
followerCount = jsonpath.jsonpath(html, '$..followerCount')  # 关注人数

name_xlsx = 'new.xlsx'
if os.path.exists(name_xlsx):
    os.remove(name_xlsx)
workbook = xlsxwriter.Workbook(name_xlsx)
sheet1 = workbook.add_worksheet('关注人数')
sheet1.write_row(0, 0, ['IP', '关注人数'])
for i in range(0, count):
    sheet1.write_row(i + 1, 0, [IP_name[i], int(followerCount[i])])
workbook.close()
data = []
for i in range(0, count):
    S_count, S_id, S_name, productCount, wishCount = getSeriesInfo(IP_id[i])
    for j in range(0,S_count):
        productcount, productidd, productnamee, minPrice, orderCount, minOnlinePrice = getProductInfo(S_id[j])
        for k in range(0, productcount):
            if minOnlinePrice[k] is None:
                if minPrice[k] is None:
                    price = 0
                else:
                    price = float(minPrice[k])
            else:
                if float(minOnlinePrice[k]) == 0:
                    if minPrice[k] is not None:
                        price = float(minPrice[k])
                    else:
                        price = 0
                else:
                    price = float(minOnlinePrice[k])
            data.append({'IP': IP_name[i],
                         '系列': S_name[j],
                         '产品': productnamee[k],
                         '价格': price,
                         '付款人数': int(orderCount[k])})
            print(data[-1])
dataframe = pd.DataFrame(data)
excelPath = name_xlsx
excelWriter = pd.ExcelWriter(excelPath, engine='openpyxl')
sheet2 = '交易信息'
excelAddSheet(dataframe, sheet2, excelWriter)


