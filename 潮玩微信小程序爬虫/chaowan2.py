from chaowan import getSeriesID, getSeriesInfo, getProductInfo, data
import pandas as pd
from openpyxl import load_workbook
import os
import xlsxwriter

# 这里改关键词
keyword = 'LABUBU'  # 自行更改关键词

r_kw = getSeriesID(keyword)
name = r_kw[0]['name']
SeriesID = r_kw[0]['id']
count, productid, productname, productCount, wishCount = getSeriesInfo(SeriesID)
name_xlsx = '{}系列信息.xlsx'.format(keyword)

# 删除重名的旧excel
if os.path.exists(name_xlsx):
    os.remove(name_xlsx)

# 创建新excel
workbook = xlsxwriter.Workbook(name_xlsx)
worksheet = workbook.add_worksheet('潮玩{}系列'.format(keyword))
worksheet.write_row(0, 0, ['名称', '在卖人数', '想要人数'])
data = data(keyword)
for i in range(0, len(data)):
    worksheet.write_row(i + 1, 0, data[i])

workbook.close()


def excelAddSheet(dataframe, sheetname, excelWriter):
    book = load_workbook(excelWriter.path)
    excelWriter.book = book
    dataframe.to_excel(excel_writer=excelWriter, sheet_name=sheetname, index=None)
    excelWriter.close()


for i in range(0, count):
    value_title = [['名称', '价格', '付款人数']]  # 表头
    productcount, productidd, productnamee, minPrice, orderCount, minOnlinePrice = getProductInfo(productid[i])
    data = []
    for j in range(0, productcount):
        if minOnlinePrice[j] is None:
            minOnlinePrice[j] = minPrice[j]
        data.append({'名称': productnamee[j], '价格': float('%.2f' % float(minOnlinePrice[j])), '付款人数': int(orderCount[j])})
    excelPath = name_xlsx
    dataframe = pd.DataFrame(data)
    excelWriter = pd.ExcelWriter(excelPath, engine='openpyxl')
    sheetname = productname[i]
    excelAddSheet(dataframe, sheetname, excelWriter)
